from pymongo import MongoClient
from openpyxl import Workbook, drawing
import subprocess
import json
import shlex
from io import BytesIO
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import requests
import argparse

# Setup MongoDB connection
client = MongoClient('mongodb://localhost:27017/')
db = client['media_database']  # Adjust the database name as necessary

def import_xytech(file_path):
    collection = db['xytech']
    data = {}
    locations = []
    with open(file_path, 'r') as file:
        lines = [line.strip() for line in file if line.strip()]
        
        # Check for sufficient data before proceeding
        if len(lines) < 8:  # Ensuring that there are at least enough lines for basic data
            raise ValueError("Insufficient data in file")

        # Safe extraction function
        def extract_data(line, expect_colon=True):
            parts = line.split(': ')
            if expect_colon and len(parts) < 2:
                raise ValueError(f"Expected ':' not found in line: {line}")
            return parts[1].strip() if len(parts) > 1 else line

        # Extract Producer, Operator, Job using extract_data expecting colon
        data['Producer'] = extract_data(lines[1])
        data['Operator'] = extract_data(lines[2])
        data['Job'] = extract_data(lines[3])
        
        # Extract Notes not necessarily expecting colon
        data['Notes'] = extract_data(lines[-1], expect_colon=False)
        
        # Assuming location data starts from a specific line and ends before the notes
        start_index = 5  # Adjust based on your actual data
        end_index = -2  # Adjust based on your actual data
        for line in lines[start_index:end_index]:
            if '/' in line:
                locations.append(line)
        
        data['Locations'] = locations
        collection.insert_one(data)

def import_baselight(file_path):
    collection = db['baselight']
    with open(file_path, 'r') as file:
        for line in file:
            if line.strip():
                parts = line.strip().split(' ')
                path = parts[0]
                frames = parts[1:]
                # Cleaning up frames to remove any non-numeric entries
                frames = [frame for frame in frames if frame.replace('<err>', '').replace('<null>', '').isnumeric()]
                collection.insert_one({'path': path, 'frames': frames})

def get_video_duration(video_path):
    """Returns the duration of the video in seconds."""
    try:
        # Construct the command to get file details
        command = [
            'ffprobe', 
            '-v', 'error',  # no verbose
            '-show_entries', 'format=duration',  # show duration
            '-of', 'default=noprint_wrappers=1:nokey=1',  # output format options
            video_path
        ]
        
        # Execute the command
        result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        
        # Get the duration
        duration = float(result.stdout)
        
        return duration
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

def frame_to_timecode(total_frames, frame_rate):
    """Convert frame count to a timecode string."""
    hours = (total_frames / frame_rate) // 3600
    minutes = ((total_frames / frame_rate) % 3600) // 60
    seconds = (total_frames / frame_rate) % 60
    frames = total_frames % frame_rate
    return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}:{int(frames):02}"

def fetch_data_from_mongo():
    client = MongoClient('mongodb://localhost:27017/')
    db = client['media_database']

    xytech_data = db.xytech.find_one()  # Assuming there's only one xytech document
    baselight_data = list(db.baselight.find())

    return xytech_data, baselight_data

def get_video_frame_rate(video_path):
    """Returns the frame rate of the video as a float."""
    try:
        # Construct the command to get video stream details
        command = [
            'ffprobe',
            '-v', 'error',  # suppress errors
            '-select_streams', 'v:0',  # select the first video stream
            '-show_entries', 'stream=r_frame_rate',  # show the frame rate
            '-of', 'default=noprint_wrappers=1:nokey=1',  # output formatting options
            video_path
        ]
        
        # Execute the command
        result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
        
        # Process the frame rate
        r_frame_rate = result.stdout.strip()
        # r_frame_rate usually comes in the form of a fraction "num/den"
        num, den = map(int, r_frame_rate.split('/'))
        frame_rate = num / den if den != 0 else num  # Calculate the frame rate as a float

        return frame_rate
    except Exception as e:
        print(f"An error occurred: {e}")
        return None
    
def create_thumbnail(video_path, start_frame, end_frame=None):
    
    middle_frame = start_frame if end_frame is None else start_frame + (end_frame - start_frame) // 2

    # Command to extract the frame and resize it to 96x74 pixels
    command = f"ffmpeg -i {video_path} -vf \"select=eq(n\,{middle_frame}),scale=96:74\" -vframes 1 -f image2pipe -"
    
    # Execute the command using subprocess
    try:
        result = subprocess.run(shlex.split(command), stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True)
        # Create an Image object from binary data
        image_data = BytesIO(result.stdout)
        img = Image(image_data)
        return img
    except subprocess.CalledProcessError as e:
        print(f"Failed to create thumbnail: {e.stderr.decode()}")
        return None

def add_row_and_image(ws, row_data, img):
    row = ws.append(row_data)
    if img:
        # Calculate the Excel cell reference where the image will be added
        cell_reference = f"{get_column_letter(len(row_data))}{ws.max_row}"
        img.anchor = cell_reference
        ws.add_image(img)

def create_excel_file(xytech_data, baselight_data, video_path, frame_rate):
    wb = Workbook()
    ws = wb.active

    # Set column widths for readability
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20

    duration = get_video_duration(video_path)
    total_frames = duration * frame_rate
    

    # Add Xytech data
    ws.append(['Producer', 'Operator', 'Job', 'Notes'])
    ws.append([
        xytech_data.get('Producer', ''),
        xytech_data.get('Operator', ''),
        xytech_data.get('Job', ''),
        xytech_data.get('Notes', '')
    ])

    # Add some space between sections
    ws.append([])
    ws.append([])

    # Add Baselight data header
    ws.append(['Location', 'Frames', 'Timecode', 'Thumbnail'])

    # Create a dictionary for mapping baselight paths to xytech locations
    xytech_locations = {loc.split('/')[-1]: loc for loc in xytech_data['Locations']}

    # Process and add Baselight data
    for item in baselight_data:
        baselight_path = item['path']
        baselight_suffix = baselight_path.split('/')[-1]

        # Find matching xytech location
        xytech_path = xytech_locations.get(baselight_suffix)
        if xytech_path:
            frames = item['frames']
            frames = sorted(map(int, frames))  # Convert all to integers and sort

            # Split frames into ranges and append each as a separate entry
            if frames:
                start = frames[0]
                end = start
                if start <= total_frames:
                    for i in range(1, len(frames)):
                        if frames[i] != frames[i-1] + 1:
                            img = create_thumbnail(video_path, (start + end) // 2)
                            if start == end:
                                ws.append([xytech_path, f"{start}", f"{frame_to_timecode(start, frame_rate)}"])
                            else:
                                ws.append([xytech_path, f"{start}-{end}", f"{frame_to_timecode(start, frame_rate)} - {frame_to_timecode(end, frame_rate)}"])
                            img.anchor = f"{get_column_letter(4)}{ws.max_row}"
                            ws.add_image(img)
                            start = frames[i]
                        end = frames[i]
                    img = create_thumbnail(video_path, (start + end) // 2)
                    if start == end:
                        ws.append([xytech_path, f"{start}", f"{frame_to_timecode(start, frame_rate)}"])
                    else:
                        ws.append([xytech_path, f"{start}-{end}", f"{frame_to_timecode(start, frame_rate)} - {frame_to_timecode(end, frame_rate)}"])
                    img.anchor = f"{get_column_letter(4)}{ws.max_row}"
                    ws.add_image(img)
                        
        else:
            print(f"No xytech match found for baselight path: {baselight_path}")

    # Save the workbook
    wb.save('combined_data.xlsx')
    print("Excel file has been saved.")

def parse_frames_to_timecode(frame_range, frame_rate):
    """Converts frame ranges into start and end times in seconds."""
    start_frame, end_frame = map(int, frame_range.split('-'))
    start_seconds = start_frame / frame_rate
    end_seconds = end_frame / frame_rate if '-' in frame_range else start_seconds
    return start_seconds, end_seconds

def extract_segment(video_path, start_seconds, end_seconds, output_file, frame_rate):
    """Extracts a segment from a video file using ffmpeg."""
    duration = end_seconds - start_seconds
    cmd = [
        'ffmpeg',
        '-ss', str(start_seconds),
        '-i', video_path,
        '-t', str(duration),
        '-c', 'copy',
        output_file
    ]
    subprocess.run(cmd, check=True)

def upload_to_frame_io(file_path, access_token, project_id):
    url = 'https://api.frame.io/v2/assets'
    headers = {'Authorization': f'Bearer {access_token}'}
    files = {'file': (file_path, open(file_path, 'rb'), 'application/octet-stream')}
    data = {'parent_asset_id': (None, project_id, 'application/json')}  # Ensure this is sent as part of the multipart form data
    response = requests.post(url, headers=headers, files=files, data=data)
    print(f"Status: {response.status_code}, Response: {response.text}")
    files['file'][1].close()  # Ensure to close the file after sending it

def parse_timecode_to_seconds(timecode, frame_rate=24):
    """
    Converts a timecode string formatted as 'HH:MM:SS:FF' into total seconds.
    
    Args:
    timecode (str): The timecode string to convert.
    frame_rate (int): The number of frames per second. Default is 24.

    Returns:
    float: Total seconds represented by the timecode.
    """
    parts = timecode.split(':')
    if len(parts) != 4:
        raise ValueError("Timecode format should be HH:MM:SS:FF")

    hours, minutes, seconds, frames = map(int, parts)
    total_seconds = hours * 3600 + minutes * 60 + seconds + frames / frame_rate
    return total_seconds


def process_and_upload_frames(video_path, frame_data, frame_rate, access_token, project_id):
    for entry in frame_data:
        frame_info, time_range = entry.split('\t')
        time_parts = time_range.split(' - ')
        
        if len(time_parts) == 1:
            start_time = end_time = time_parts[0]
        else:
            start_time, end_time = time_parts

        # Calculate start and end seconds for ffmpeg
        start_seconds = parse_timecode_to_seconds(start_time)
        end_seconds = parse_timecode_to_seconds(end_time) if start_time != end_time else start_seconds + 1 / frame_rate

        output_file = f"segment_{start_seconds}_{end_seconds}.mp4"
        extract_segment(video_path, start_seconds, end_seconds, output_file, frame_rate)
        upload_to_frame_io(output_file, access_token, project_id)


def main():
    frame_rate = 0
    video_path = ""
    frame_data = []

    parser = argparse.ArgumentParser(description="Projec 3")
    parser.add_argument("--baselight", type=str, help="path to baselight file")
    parser.add_argument("--xytech", type=str, help="path to xytecht file")
    parser.add_argument("--process", type=str, help="path to video file")
    parser.add_argument("--output", type=str, help="outputs xls and frame io request")

    args = parser.parse_args()

    if args.baselight:
        import_baselight(args.baselight)
    if args.xytech:
        import_xytech(args.xytech)
    if args.process:
        frame_rate = get_video_frame_rate(args.process)
        video_path = args.process
    if args.output:
        token = "fio-u-GizuPM53J7La7onLmcxNqe5KAJ6VJ4jxsOjEe6v0ROaYFVWSlgM0M7Vaw4d8krOn"
        project_id = "13702e16-e0c9-4ed3-b2d1-30b8b36eb9a5"
        xytech_data, baselight_data = fetch_data_from_mongo()
        create_excel_file(xytech_data, baselight_data, video_path, frame_rate)
        process_and_upload_frames(video_path, frame_data, frame_rate, token, project_id)
    

if __name__ == "__main__":
    main()
